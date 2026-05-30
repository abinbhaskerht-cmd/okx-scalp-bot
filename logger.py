# ============================
#   TRADE LOGGER
# ============================

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

LOG_FILE = 'trade_log.xlsx'


def create_log_file():
    if os.path.exists(LOG_FILE):
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trade Log'

    headers = [
        'Date', 'Time', 'Symbol', 'Signal',
        'Entry Price', 'Quantity', 'Stop Loss',
        'Take Profit', 'Order ID', 'Status'
    ]

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')

    widths = [12, 10, 12, 8, 12, 12, 12, 12, 25, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    wb.save(LOG_FILE)
    print(f"   📊 Log file created: {LOG_FILE}")


def log_trade(symbol, signal, entry_price, qty, sl, tp, order_id):
    create_log_file()

    wb   = openpyxl.load_workbook(LOG_FILE)
    ws   = wb.active
    now  = datetime.now()
    date = now.strftime('%Y-%m-%d')
    time = now.strftime('%H:%M:%S')

    if signal == 'buy':
        fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    else:
        fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

    row_data = [
        date, time, symbol, signal.upper(),
        entry_price, qty, sl, tp, order_id, 'OPEN'
    ]

    next_row = ws.max_row + 1
    for col, value in enumerate(row_data, 1):
        cell           = ws.cell(row=next_row, column=col, value=value)
        cell.fill      = fill
        cell.alignment = Alignment(horizontal='center')

    wb.save(LOG_FILE)
    print(f"   📊 Trade logged to {LOG_FILE}")


def get_trade_summary():
    if not os.path.exists(LOG_FILE):
        print("   No trades logged yet.")
        return

    wb    = openpyxl.load_workbook(LOG_FILE)
    ws    = wb.active
    total = ws.max_row - 1
    buys  = 0
    sells = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] == 'BUY':
            buys += 1
        elif row[3] == 'SELL':
            sells += 1

    print(f"\n   📊 Trade Summary")
    print(f"   Total Trades : {total}")
    print(f"   Buy Trades   : {buys}")
    print(f"   Sell Trades  : {sells}")