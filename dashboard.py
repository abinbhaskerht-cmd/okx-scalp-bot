# ============================
#   PERFORMANCE DASHBOARD
# ============================

import os
import openpyxl
from datetime import datetime

LOG_FILE = 'trade_log.xlsx'


def get_stats():
    """Calculate performance statistics"""
    if not os.path.exists(LOG_FILE):
        return None

    wb  = openpyxl.load_workbook(LOG_FILE)
    ws  = wb.active

    if ws.max_row < 2:
        return None

    trades     = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            trades.append({
                'date'   : row[0],
                'time'   : row[1],
                'symbol' : row[2],
                'signal' : row[3],
                'entry'  : row[4],
                'qty'    : row[5],
                'sl'     : row[6],
                'tp'     : row[7],
                'status' : row[9],
            })

    total  = len(trades)
    wins   = len([t for t in trades if t['status'] == 'WIN'])
    losses = len([t for t in trades if t['status'] == 'LOSS'])
    open_t = len([t for t in trades if t['status'] == 'OPEN'])

    win_rate = (wins / (wins + losses) * 100) if (wins + losses) > 0 else 0

    return {
        'total'   : total,
        'wins'    : wins,
        'losses'  : losses,
        'open'    : open_t,
        'win_rate': win_rate,
        'trades'  : trades
    }


def print_dashboard():
    """Print performance dashboard in terminal"""
    stats = get_stats()

    print("\n" + "=" * 45)
    print("   📊 PERFORMANCE DASHBOARD")
    print("=" * 45)

    if not stats:
        print("   No trades logged yet.")
        print("=" * 45)
        return

    print(f"\n   📈 TRADE SUMMARY")
    print(f"   {'─'*35}")
    print(f"   Total Trades  : {stats['total']}")
    print(f"   Wins          : {stats['wins']} ✅")
    print(f"   Losses        : {stats['losses']} ❌")
    print(f"   Open          : {stats['open']} ⏳")
    print(f"   Win Rate      : {stats['win_rate']:.1f}%")
    print(f"   {'─'*35}")

    if stats['win_rate'] >= 50:
        print(f"   Status: ✅ PROFITABLE STRATEGY")
    elif stats['win_rate'] >= 40:
        print(f"   Status: ⚠️  MARGINAL STRATEGY")
    else:
        print(f"   Status: ❌ NEEDS IMPROVEMENT")

    print(f"\n   📋 RECENT TRADES")
    print(f"   {'─'*35}")
    recent = stats['trades'][-5:] if len(stats['trades']) >= 5 else stats['trades']
    for t in reversed(recent):
        emoji = "✅" if t['status'] == 'WIN' else "❌" if t['status'] == 'LOSS' else "⏳"
        print(f"   {emoji} {t['date']} {t['time']} | {t['symbol']} | {t['signal']}")

    print("\n" + "=" * 45)


def get_telegram_summary():
    """Get dashboard summary for Telegram"""
    stats = get_stats()

    if not stats:
        return "📊 No trades logged yet."

    msg = f"""📊 <b>PERFORMANCE DASHBOARD</b>

📈 <b>Trade Summary</b>
Total Trades : <code>{stats['total']}</code>
Wins         : <code>{stats['wins']}</code> ✅
Losses       : <code>{stats['losses']}</code> ❌
Open         : <code>{stats['open']}</code> ⏳
Win Rate     : <code>{stats['win_rate']:.1f}%</code>"""

    if stats['win_rate'] >= 50:
        msg += "\n\n✅ <b>Strategy is PROFITABLE</b>"
    elif stats['win_rate'] >= 40:
        msg += "\n\n⚠️ <b>Strategy is MARGINAL</b>"
    else:
        msg += "\n\n❌ <b>Strategy needs improvement</b>"

    return msg